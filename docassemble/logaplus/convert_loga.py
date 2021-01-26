#import source file LOGA .txt File
import pandas as pd
import pandas
from openpyxl import load_workbook
import numpy as np
from io import StringIO
import re
import chardet
from docassemble.base.util import DAFile
from docassemble.base.util import path_and_mimetype

#in_folder = 'Daten/'
#in_file_name = 'lohn 2020-06'
#in_file_format = '.txt'

#to_file_format = '.csv'
#to_file_name = 'Primanota ' + in_file_name
#to_folder = 'Auswertungen/'
#template_file = r'Primanota_Template.xlsx'

#file_name_with_path = in_folder + in_file_name + in_file_format

def read_loga(the_file):
  try:
    output = StringIO()  # stringstream for import

    raw_file = open(the_file.path(), 'rb').read()
    guess_encoding = chardet.detect(raw_file)

    p = re.compile('^\"(.*)\"$')

 
    for line in open(the_file.path(), 'rt', encoding=guess_encoding['encoding']):   # the_file.readlines():
      this = p.match(line)
      if this:
        output.write(line)

    output.seek(0)  # resetz pos of stream 
    data = pd.read_csv(output, sep=";", decimal=",", header=None, encoding='utf-8',quotechar='"')
    output.close()
    error_code = 0

  except:
    error_code = -1
  
  if not error_code:
    
    dat=data.iloc[:,0].str.split(';',expand=True)
    beleg_datum = pd.to_datetime(dat.iloc[1,11])

    # excerpt relevant column and clean data types
    dat01 = dat.loc[dat[0].str.contains('D1'),[17,18,19,20,21,22,23,32,24,25]]
    dat01.columns = dat.iloc[2,1:11]
    dat01.index.name = 'Nr'
    dat01.iloc[:,7]=pd.to_datetime(dat01.iloc[:,7])
    dat01.iloc[:,8]=dat01.iloc[:,8].str.replace(',','.').astype(float)
    dat01.iloc[:,9]=dat01.iloc[:,9].str.replace(',','.').astype(float)

    # take just rows with cost-no.
    dat01 = dat01.loc[dat01['Kostenart'].astype(str).str.isdigit(),:]
  
    dat01.loc[dat01['Betrag SOLL']>0, 'SOLL/Haben-Kz']="S"
    dat01.loc[dat01['Betrag Haben']>0, 'SOLL/Haben-Kz']="H"
    dat01['Umsatz']=dat01.loc[:,['Betrag SOLL','Betrag Haben']].max(axis=1)
    dat01['Gegenkonto']=175500
    dat01['Belegfeld 1'] = the_file[0].filename    #in_file_name + in_file_format
    dat01['Konto']=dat01.apply(lambda x: pd.to_numeric(x['Kostenart'] if len(x['Kostenart'])>4 else x['Kostenart']+"00"), axis=1)
    dat01['Buchungstext']= dat01.apply(lambda x: " ".join([x['Kostenart-Bezeichnung'],x['zu belastende Kostenstelle'],x['Bezeichn. / MwSt Schl.'],'Z-Datum',x['Zuordn. Datum'].strftime('%d.%m.%Y')]), axis=1 )
    dat01['Kostenstelle']= pd.to_numeric(dat01['Kostenstelle'])
    dat01['Belegdatum'] = beleg_datum
  

    (template_file,mimetype) = path_and_mimetype('data/sources/Primanota_Template.xlsx')
    book = load_workbook(template_file)
    ws = book.active
    template_cols = []
    for c in ws['3']:
      template_cols.append(c.value)
    template_df = pd.DataFrame({'Umsatz (ohne Soll/Haben-Kz)': dat01['Umsatz'],
                            'Soll/Haben-Kennzeichen': dat01['SOLL/Haben-Kz'],
                            'Konto': dat01['Konto'],
                            'Gegenkonto (ohne BU-Schlüssel)': dat01['Gegenkonto'],
                            'Belegdatum': dat01['Belegdatum'],
                            'Belegfeld 1': dat01['Belegfeld 1'],
                            'Buchungstext': dat01['Buchungstext'],                            
                            'KOST1 - Kostenstelle': dat01['Kostenstelle']
                           },columns=template_cols).astype({'Umsatz (ohne Soll/Haben-Kz)':'float','Belegdatum':'datetime64[ns]','KOST1 - Kostenstelle':'Int64'}) 

    test_data = pd.DataFrame(pd.pivot_table(dat01,index="SOLL/Haben-Kz",values="Umsatz",aggfunc='sum').to_records())
    myfile =  DAFile()
    myfile.initialize(filename=the_file[0].filename.split('.')[0]+".csv")
    template_df.to_csv(myfile.path(), sep=';',index=False,header=True,decimal=',',date_format='%d.%m.%Y',encoding='latin-1')

    return {'error_code':error_code, 'date':beleg_datum, 'test':test_data, 'file':myfile}
  else:
    return {'error_code':error_code, 'date':None, 'test':pd.DataFrame(), 'file':DAFile()}